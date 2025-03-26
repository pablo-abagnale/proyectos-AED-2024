from openpyxl import workbook
from openpyxl.utils import get_column_letter
import openpyxl
from Class_3 import *
import datetime
import os
import platform
import calendar
from openpyxl.worksheet.table import Table, TableStyleInfo

wb_nomina = openpyxl.load_workbook('Nomina_Omnia.xlsx')  # Abrir la nomina de Omnia
st_nomina = wb_nomina['nomina']  # Trabajar sobre la hoja Nomina


def clear_console():
    if platform.system() == "Windows":
        os.system("cls")
    else:
        os.system("clear")

def ingresar_mes():
    m = int(input("Ingrese un mes: "))
    while m <= 0 or m > 12:
        m = int(input("Ingrese un número válido (1-12): "))

    y = datetime.datetime.now().year
    num_days = calendar.monthrange(y, m)[1]

    days_list = [datetime.date(y, m, day).strftime("%d/%m/%Y") for day in range(1, num_days + 1)]

    return days_list


def cargar():

    wb_nomina = openpyxl.load_workbook('Nomina_Omnia.xlsx')  # Abrir la nomina de Omnia
    st_nomina = wb_nomina['nomina']  # Trabajar sobre la hoja Nomina

    array_agentes = []

    for row in st_nomina.iter_rows(min_row=2, values_only=True):
        if row[27] is not None:
            legajo = int(row[0])
            documento = int(row[3])
            agente = f"{row[1]} {row[2]}"
            mail = row[27]
            supervisor = row[29]
            cuenta = row[16]

            agente_obj = Agente(legajo, documento, agente, mail, supervisor, cuenta)
            array_agentes.append(agente_obj)

    wb_nomina.close()

    return array_agentes

def insertar_ordenado(array, dato):#Guardar datos ordenados en el control de horas
    n = len(array)
    izq = 0
    der = n - 1
    while izq <= der:
        centro = (izq + der) // 2
        if array[centro].fecha == dato.fecha:
            pos = centro
            break
        elif array[centro].fecha > dato.fecha:
            der = centro - 1
        else:
            izq = centro + 1
    else:
        pos = izq

    array[pos:pos] = [dato]

def mostrar(array): # Mostrar arrays
    for i in array:
        print('-' * 200)
        print(i)

def pedir_fecha(fecha_str):
    while True:
        fecha_str
        try:
            #fecha = datetime.strptime(fecha_str, "%d/%m/%Y")
            fecha = datetime.datetime.strptime(fecha_str, '%d/%m/%Y')
            fecha_formateada = fecha.strftime("%d/%m/%Y")
            return fecha_formateada
        except ValueError:
            print("Formato de fecha incorrecto. Inténtelo de nuevo.")
def obtener_datos_antes_de_coma(celda):
    # Encuentra la posición de la coma
    posicion_coma = celda.find(',')
    if posicion_coma != -1:
        # Extrae el texto antes de la coma
        return celda[:posicion_coma]
    else:
        return celda
def shell_short(v):
    n = len(v)
    h = 1
    while h <= n // 9:
        h = 3 * h + 1
    while h > 0:
        for j in range(h, n):
            y = v[j].fecha
            k = j - h
            while k >= 0 and y < v[k].fecha:
                v[k + h].fecha = v[k].fecha
                k -= h
            v[k + h].fecha = y
        h //= 3
    return v



def convertir_a_horas(minutos):
        horas = minutos // 60
        minutos_restantes = minutos % 60
        return f"{int(horas):02d}:{int(minutos_restantes):02d}"

def sumador_horas(array, servicio, mes):
        vueltas = []

        for dia in mes:
            h_productivas = 0
            t_break = 0
            t_ocupado = 0
            t_almuerzo = 0
            t_coach = 0
            dato = False

            for i in range(len(array)):
                fecha = array[i].fecha
                if fecha == dia and array[i].cola == servicio:
                    dato = True
                    if array[i].estado == 'En línea':
                        h_productivas += array[i].time
                    elif array[i].estado == 'Break':
                        t_break += array[i].time
                    elif array[i].estado == 'Ocupado':
                        t_ocupado += array[i].time
                    elif array[i].estado == 'Almuerzo':
                        t_almuerzo += array[i].time
                    elif array[i].estado == 'Coach':
                        t_coach += array[i].time

            if dato:
                vuelta = (
                    dia, servicio,'HS Productivas',
                    convertir_a_horas(h_productivas),'Tiempo en break',
                    convertir_a_horas(t_break),'Tiempo en ocupado',
                    convertir_a_horas(t_ocupado),'Tiempo en Coaching',
                    convertir_a_horas(t_coach),'Tiempo en almuerzo',
                    convertir_a_horas(t_almuerzo)
                )
                vueltas.append(vuelta)

        mostrar(vueltas)

def guardar_cntl_horas(array):

    archivo = 'ctrl_hs_digital.xlsx'
    try:
        wb = openpyxl.load_workbook(archivo)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    if 'BBDD_HS' in wb.sheetnames:
        ws_BBDD = wb['BBDD_HS']
    else:
        ws_BBDD = wb.create_sheet('BBDD_HS')

    # Verificar si la tabla "cntrl" existe
    tabla_existente = None
    for tbl in ws_BBDD.tables.values():
        if tbl.name == "cntrl":
            tabla_existente = tbl
            break

    if tabla_existente:
        # Borrar todos los datos de la tabla existente
        ws_BBDD.delete_rows(2, ws_BBDD.max_row)
    else:
        # Añadir encabezados si la tabla no existe
        ws_BBDD.append(['Mail', 'Agente', 'Fecha', 'Servicio', 'Estado', 'Duracion'])

    # Añadir datos
    for i in array:
        ws_BBDD.append([i.mail, i.agente, i.fecha, i.cola, i.estado, i.time])

    # Definir el rango de la tabla
    min_col = 1
    max_col = ws_BBDD.max_column
    min_row = 1
    max_row = ws_BBDD.max_row
    tabla_rango = f"A{min_row}:F{max_row}"

    if tabla_existente:
        # Actualizar el rango de la tabla existente
        tabla_existente.ref = tabla_rango
    else:
        # Crear Tabla
        tabla = Table(displayName="cntrl", ref=tabla_rango)
        estilo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        tabla.tableStyleInfo = estilo
        ws_BBDD.add_table(tabla)

    wb.save(archivo)
    wb.close()


def ctrl_horas(v):
    #mes = ingresar_mes()
    wb_botmaker = openpyxl.load_workbook('botmaker.xlsx')
    st_botmaker = wb_botmaker['Sheet1']
    array_valores = []
    gxf = ['olga.reggiori@naranja.com','gaston.areco@naranja.com','ana.salomonpersico@naranja.com','tomas.moyano@naranjax.com','micaela.ceballos@naranjax.com']
    q = 0

    for i in v:   #Valida que solo se traigan los datos de agentes EVOLTIS
        m = i.mail
        for row in st_botmaker.iter_rows(min_row=2, max_row=st_botmaker.max_row, min_col=0, max_col=st_botmaker.max_column):

            if row[0].value == m:
                q += 1
                if row[2].value == 'Asesor' or row[2].value == 'Supervisores CC':
                    mail = row[0].value
                    agente = row[1].value
                    rol = row[2].value
                    if row[2].value =='Supervisores CC' and row[0].value in gxf:
                        cola = 'GxF'
                    elif row[2].value =='Supervisores CC' and row[0].value  not in gxf:
                        cola = 'Super'
                    elif 'FacebookMuro'in row[3].value or 'InstagramFeed'in row[3].value:
                        cola = 'RRSS'
                    elif 'Onboarding'in row[3].value or 'piloto'in row[3].value or 'MGX'in row[3].value:
                        cola = 'FCR'
                    elif 'mailTarjetaDeCredito'in row[3].value or 'mailFintech'in row[3].value:
                        cola = 'Mail'
                    elif 'ChatInApp'in row[3].value or 'WebChat'in row[3].value:
                        cola = 'ChatInApp'
                    else:
                        cola = row[3].value
                    if 'En línea (no disponible)'in row[4].value:
                        estado = 'En línea'
                    else:
                        estado = row[4].value
                    fecha = obtener_datos_antes_de_coma(row[6].value)
                    fecha = pedir_fecha(fecha)
                    time = row[8].value/60
                    if cola != 'Super':
                        valores = Control(mail, agente, rol, cola, estado, fecha, time)
                        array_valores.append(valores)
                        #insertar_ordenado(array_valores,valores)

    print('\n',q,'\n')

    guardar_cntl_horas(array_valores)
