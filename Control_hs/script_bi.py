from openpyxl import workbook
from openpyxl.utils import get_column_letter
import openpyxl
from objetos import *
import datetime
import os
import platform
import calendar
import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo

class Agente:
    def __init__(self, legajo,documento,agente,mail,supervisor,cuenta):
        self.legajo = legajo  #Legajo OMNIA
        self.documento = documento #DNI sin puntos
        self.agente = agente # Nombre completo Apellido + Nombre
        self.mail = mail #Mail Naranja NX / Usuario Botmaker
        self.supervisor = supervisor #Equipo
        self.cuenta = cuenta #Sub campaña

    def __str__(self):
        return f'DNI: {self.documento:<10} Agente: {self.agente:<35} mail: {self.mail:<40} Equipo : {self.supervisor}'

class Control:
    def __init__(self,mail,agente,rol,cola,estado,fecha,time):
        self.mail = mail
        self.agente = agente
        self.rol = rol
        self.cola = cola
        self.estado = estado
        self.fecha = fecha
        self.time = time

    def __str__(self):
        return f'mail:{self.mail:<40} rol: {self.rol:<40} servicio: {self.cola:<17} estado: {self.estado:<15} fecha: {self.fecha:<15} duracion: {self.time:<5} horas'

class Utilizacion:
    def __init__(self,mail,agente,equipo,fecha,horas,disponible,no_disponible,descanso,ocupado,coaching,almuerzo):
        self.mail = mail
        self.agente = agente
        self.equipo = equipo
        self.fecha = fecha
        self.horas = horas
        self.disponible = disponible
        self.no_disponible = no_disponible
        self.descanso = descanso
        self.ocupado = ocupado
        self.coaching = coaching
        self.almuerzo = almuerzo

    def __str__(self):
        return f'Fecha: {self.fecha} Mail:{self.mail:<30} rol: {self.agente:<30} : {self.equipo:<30} Horas: {self.horas:<10} hs Aux: {self.no_disponible + self.descanso + self.ocupado:<15}'


wb_nomina = openpyxl.load_workbook('Nomina_Omnia.xlsx')  # Abrir la nomina de Omnia
st_nomina = wb_nomina['nomina']  # Trabajar sobre la hoja Nomina


def clear_console():
    if platform.system() == "Windows":
        os.system("cls")
    else:
        os.system("clear")

def ingresar_mes():
    m = int(input("Ingrese un mes: "))
    while m <= 0 or m > 12:
        m = int(input("Ingrese un número válido (1-12): "))

    y = datetime.datetime.now().year
    num_days = calendar.monthrange(y, m)[1]

    days_list = [datetime.date(y, m, day).strftime("%d/%m/%Y") for day in range(1, num_days + 1)]

    return days_list


def cargar():

    wb_nomina = openpyxl.load_workbook('Nomina_Omnia.xlsx')  # Abrir la nomina de Omnia
    st_nomina = wb_nomina['nomina']  # Trabajar sobre la hoja Nomina

    array_agentes = []
    nomina = []
    for row in st_nomina.iter_rows(min_row=2, values_only=True):
        if row[27] is not None:
            legajo = int(row[0])
            documento = int(row[3])
            agente = f"{row[1]} {row[2]}"
            mail = row[27]
            supervisor = row[29]
            cuenta = row[16]


            row = {'Legajo': legajo,
                   'Documento': documento,
                   'Agente': agente,
                   'Mail': mail,
                   'Equipo': supervisor,
                   'cuenta': cuenta,
                   }
            nomina.append(row)

            df_nomina = pd.DataFrame(nomina)

            agente_obj = Agente(legajo, documento, agente, mail, supervisor, cuenta)
            array_agentes.append(agente_obj)

    wb_nomina.close()

    df_nomina.to_excel('Cntl_hs.xlsx', index=False, sheet_name='Nomina')

    return array_agentes

def insertar_ordenado(array, dato):#Guardar datos ordenados en el control de horas
    n = len(array)
    izq = 0
    der = n - 1
    while izq <= der:
        centro = (izq + der) // 2
        if array[centro].fecha == dato.fecha:
            pos = centro
            break
        elif array[centro].fecha > dato.fecha:
            der = centro - 1
        else:
            izq = centro + 1
    else:
        pos = izq

    array[pos:pos] = [dato]


def pedir_fecha(fecha_str):
    while True:
        fecha_str
        try:
            #fecha = datetime.strptime(fecha_str, "%d/%m/%Y")
            fecha = datetime.datetime.strptime(fecha_str, '%d/%m/%Y')
            fecha_formateada = fecha.strftime("%d/%m/%Y")
            return fecha_formateada
        except ValueError:
            break
def obtener_datos_antes_de_coma(celda):
    # Encuentra la posición de la coma
    posicion_coma = celda.find(',')
    if posicion_coma != -1:
        # Extrae el texto antes de la coma
        return celda[:posicion_coma]
    else:
        return celda
def shell_short(v):
    n = len(v)
    h = 1
    while h <= n // 9:
        h = 3 * h + 1
    while h > 0:
        for j in range(h, n):
            y = v[j].fecha
            k = j - h
            while k >= 0 and y < v[k].fecha:
                v[k + h].fecha = v[k].fecha
                k -= h
            v[k + h].fecha = y
        h //= 3
    return v

def convertir_a_horas(minutos):
        horas = minutos // 60
        minutos_restantes = minutos % 60
        return f"{int(horas):02d}:{int(minutos_restantes):02d}"

def ctrl_horas(v):
    #mes = ingresar_mes()
    wb_botmaker = openpyxl.load_workbook('botmaker.xlsx')
    st_botmaker = wb_botmaker['Sheet1']
    array_valores = []
    tabla = []
    gxf = ['olga.reggiori@naranja.com','gaston.areco@naranja.com','ana.salomonpersico@naranja.com','tomas.moyano@naranjax.com','micaela.ceballos@naranjax.com']
    q = 0

    for i in v:   #Valida que solo se traigan los datos de agentes EVOLTIS
        m = i.mail
        for row in st_botmaker.iter_rows(min_row=2, max_row=st_botmaker.max_row, min_col=0, max_col=st_botmaker.max_column):

            if row[0].value == m:
                q += 1
                if row[2].value == 'Asesor' or row[2].value == 'Supervisores CC':
                    mail = row[0].value
                    agente = row[1].value
                    rol = row[2].value
                    if row[2].value =='Supervisores CC' and row[0].value in gxf:
                        cola = 'GxF'
                    elif row[2].value =='Supervisores CC' and row[0].value  not in gxf:
                        cola = 'Super'
                    elif 'FacebookMuro'in row[3].value or 'InstagramFeed'in row[3].value:
                        cola = 'RRSS'
                    elif 'Onboarding'in row[3].value or 'piloto'in row[3].value or 'MGX'in row[3].value:
                        cola = 'FCR'
                    elif 'mailTarjetaDeCredito'in row[3].value or 'mailFintech'in row[3].value:
                        cola = 'Mail'
                    elif 'ChatInApp'in row[3].value or 'WebChat'in row[3].value:
                        cola = 'ChatInApp'
                    else:
                        cola = row[3].value
                    if 'En línea (no disponible)'in row[4].value:
                        estado = 'no disponible'
                    else:
                        estado = row[4].value
                    fecha = obtener_datos_antes_de_coma(row[6].value)
                    fecha = pedir_fecha(fecha)
                    time = row[8].value/60
                    if cola != 'Super':
                        valores = Control(mail, agente, rol, cola, estado, fecha, time)
                        insertar_ordenado(array_valores,valores)
                        row = {'Mail':mail,
                                 'Agente':agente,
                                 'Rol':rol,
                                 'Cola':cola,
                                 'Estado':estado,
                                 'Fecha':fecha,
                                 'Duracion':time
                                 }
                        tabla.append(row)

                        df = pd.DataFrame(tabla)

    return df

def utilizacion (mes):


    wb_cntrl = openpyxl.load_workbook('Cntl_hs.xlsx')  # Abrir la nomina de Omnia
    st_nomina = wb_cntrl['Nomina']  # Trabajar sobre la hoja Nomina
    st_BBDD = wb_cntrl['BBDD_HS']

    uti =[]

    for dia in mes:
        for row_i in st_nomina.iter_rows(min_row=2, max_row=st_nomina.max_row, min_col=0, max_col=st_nomina.max_column):

            mail = row_i[3].value
            horas = 0
            no_disponible = 0
            descanso = 0
            ocupado = 0
            coaching = 0
            almurzo = 0

            for row_j in st_BBDD.iter_rows(min_row=2, max_row=st_BBDD.max_row, min_col=0, max_col=st_BBDD.max_column):
                if dia == row_j[5].value:
                    if row_j[0].value == mail:

                        fecha = row_j[5].value
                        mail = row_j[0].value
                        agente = row_i[2].value
                        equipo = row_i[4].value
                        horas += row_j[6].value

                        if row_j[4].value == 'Ocupado':
                            ocupado += row_j[6].value
                        elif row_j[4].value == 'Break':
                            descanso += row_j[6].value
                        elif row_j[4].value == 'Coaching':
                            coaching += row_j[6].value
                        elif row_j[4].value == 'Almurzo':
                            almurzo += row_j[6].value
                        elif row_j[4].value == 'no disponible':
                            no_disponible += row_j[6].value

                        row = {'Fecha': fecha,
                               'Mail': mail,
                               'Agente': agente,
                               'Equipo': equipo,
                               'Jornada': horas,
                               'No disponible': no_disponible,
                               'Break': descanso,
                               'Coaching': coaching,
                               'Ocupado': ocupado,
                               'Almurzo': almurzo
                               }
                        uti.append(row)

                        df_uti = pd.DataFrame(uti)

array_agentes = cargar()
ctrl_horas(array_agentes)
